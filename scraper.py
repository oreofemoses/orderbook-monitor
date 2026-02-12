import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import time
import os
import re
import json
import requests
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- Configuration ---
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
TELEGRAM_CHAT_ID = os.getenv('TELEGRAM_CHAT_ID')

# Nigerian timezone (UTC+1)
NIGERIAN_TZ = timezone(timedelta(hours=1))

# Files & Folders
DATA_DIR = "data"
STATE_FILE = os.path.join(DATA_DIR, "health_state.json")
BASE_URL = "https://pro.quidax.io/en_US/trade/"

# Thresholds (Matching your dashboard logic)
ALERT_THRESHOLD_CYCLES = 3
ALERT_COOLDOWN_MINUTES = 30
MAX_ATTEMPTS_PER_PAIR = 3

PAIRS = [
    ['AAVE_USDT', 0.30], ['ADA_USDT', 0.26], ['ALGO_USDT', 2.00],
    ['BCH_USDT', 0.26], ['BNB_USDT', 0.30], ['BONK_USDT', 2.00],
    ['BTC_USDT', 0.20], ['CAKE_USDT', 0.30], ['CFX_USDT', 2.00],['DASH_USDT', 2.00],
    ['DOT_USDT', 0.26], ['DOGE_USDT', 0.26], ['ETH_USDT', 0.25],
    ['FARTCOIN_USDT', 2.00], ['FLOKI_USDT', 0.50], ['HYPE_USDT', 2.00],
    ['LINK_USDT', 0.26],['LSK_USDT', 1.50], ['LTC_USDT', 0.30], ['NEAR_USDT', 2.00], ['NOS_USDT', 2.00],
    ['PEPE_USDT', 0.50], ['POL_USDT', 0.50], ['QDX_USDT', 10.00],
    ['RENDER_USDT', 2.00], ['Sonic_USDT', 2.00], ['SHIB_USDT', 0.40],
    ['SLP_USDT', 2.00], ['SOL_USDT', 0.25], ['STRK_USDT', 2.00],
    ['SUI_USDT', 2.00], ['TON_USDT', 0.30], ['TRX_USDT', 0.30],
    ['USDC_USDT', 0.02], ['WIF_USDT', 2.00], ['XLM_USDT', 0.30],
    ['XRP_USDT', 0.30], ['XYO_USDT', 1.00], ['ZKSync_USDT', 2.00],
    ['BTC_NGN', 0.50], ['USDT_NGN', 0.52], ['QDX_NGN', 10.00],
    ['ETH_NGN', 0.50], ['TRX_NGN', 0.50], ['XRP_NGN', 0.50],
    ['DASH_NGN', 0.50], ['LTC_NGN', 0.50], ['SOL_NGN', 0.50],
    ['USDC_NGN', 0.50]
]

# --- Helper function to get current Nigerian time ---
def get_nigerian_time():
    """Returns current time in Nigerian timezone (UTC+1)"""
    return datetime.now(NIGERIAN_TZ)

# --- Core Computation Functions ---

def parse_number(value):
    if not value or "--" in str(value): return None
    try:
        val_str = str(value).replace(',', '').strip()
        if '{' in val_str:
            match = re.search(r"0\.0\{(\d+)\}(\d+)", val_str)
            if match:
                val_str = "0." + ("0" * int(match.group(1))) + match.group(2)
        if val_str.upper().endswith('K'): return float(val_str[:-1]) * 1_000
        if val_str.upper().endswith('M'): return float(val_str[:-1]) * 1_000_000
        return float(val_str)
    except: return None

def parse_orderbook(text: str):
    lines = text.split("\n")
    asks, bids, side, spread_pct = [], [], "asks", None
    for line in lines:
        if "Spread" in line:
            side = "bids"
            continue
        parts = line.split()
        if len(parts) == 2 and "(" in parts[1]:
            try: spread_pct = float(parts[1].replace('(', '').replace('%)', '').replace('+', ''))
            except: pass
        elif len(parts) == 3:
            p, a, t = parse_number(parts[0]), parse_number(parts[1]), parse_number(parts[2])
            if a is not None:
                row = {"price": p, "amount": a, "total": t}
                if side == "asks": asks.append(row)
                else: bids.append(row)
    return pd.DataFrame(asks), pd.DataFrame(bids), spread_pct

def calculate_liquidity_depth(asks_df, bids_df, spread_pct_range):
    if asks_df.empty or bids_df.empty: return 0
    mid = (asks_df['price'].min() + bids_df['price'].max()) / 2
    upper, lower = mid * (1 + spread_pct_range / 100), mid * (1 - spread_pct_range / 100)
    bid_depth = (bids_df[bids_df['price'] >= lower]['price'] * bids_df[bids_df['price'] >= lower]['amount']).sum()
    ask_depth = (asks_df[asks_df['price'] <= upper]['price'] * asks_df[asks_df['price'] <= upper]['amount']).sum()
    return bid_depth + ask_depth

def calculate_dws(asks_df, bids_df, num_levels=10):
    if asks_df.empty or bids_df.empty: return 0
    mid = (asks_df['price'].min() + bids_df['price'].max()) / 2
    a_sub, b_sub = asks_df.nsmallest(num_levels, 'price'), bids_df.nlargest(num_levels, 'price')
    num = (a_sub['amount'] * (a_sub['price'] - mid)).abs().sum() + (b_sub['amount'] * (mid - b_sub['price'])).abs().sum()
    den = a_sub['amount'].sum() + b_sub['amount'].sum()
    return (num / den) / mid * 100 if den > 0 else 0

def format_depth(val):
    if not val: return "$0"
    if val >= 1_000_000: return f"${val/1_000_000:.2f}M"
    if val >= 1_000: return f"${val/1_000:.1f}K"
    return f"${val:.0f}"

# --- Persistence & Helpers ---

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r') as f: return json.load(f)
    return {}

def save_state(state):
    with open(STATE_FILE, 'w') as f: json.dump(state, f)

def send_telegram(msg):
    if not TELEGRAM_BOT_TOKEN: return
    requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage", 
                 json={'chat_id': TELEGRAM_CHAT_ID, 'text': msg, 'parse_mode': 'HTML'})

def init_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
    chrome_options.add_argument(f"user-agent={user_agent}")
    if os.path.exists("/usr/bin/chromium-browser"): chrome_options.binary_location = "/usr/bin/chromium-browser"
    service = Service("/usr/bin/chromedriver")
    try: return webdriver.Chrome(service=service, options=chrome_options)
    except: return webdriver.Chrome(options=chrome_options)

# --- Excel Logging Functions ---

def get_daily_log_path():
    """Get path for today's Excel log file"""
    nigerian_time = get_nigerian_time()
    today = nigerian_time.strftime("%Y-%m-%d")
    return os.path.join(DATA_DIR, f"daily_log_{today}.xlsx")

def initialize_daily_log(log_path):
    """Create new Excel file with initial structure"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Health"
    
    # Headers
    ws['A1'] = 'Market'
    ws['B1'] = '% Spd'
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    
    # Add all markets
    for idx, (symbol, target) in enumerate(PAIRS, start=2):
        ws[f'A{idx}'] = symbol
        ws[f'B{idx}'] = f"{target}%"
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 8
    
    wb.save(log_path)
    return wb

def update_daily_log(results_dict):
    """
    Update Excel log with new check results
    results_dict: {symbol: {'status': 'CHECKED'/'WARNING'/'SKIPPED', 'depth_1.25x': val, 'depth_1.5x': val}}
    """
    log_path = get_daily_log_path()
    nigerian_time = get_nigerian_time()
    check_time = nigerian_time.strftime("%H:%M:%S")
    
    # Load or create workbook
    if os.path.exists(log_path):
        wb = load_workbook(log_path)
        ws = wb.active
    else:
        wb = initialize_daily_log(log_path)
        ws = wb.active
    
    # Find the next available column (after Market and % Spd)
    # Check if there's already a Depth column at the end
    last_col = ws.max_column
    depth_col_idx = None
    
    # Check if last column is "Depth"
    if ws.cell(1, last_col).value and "Depth" in str(ws.cell(1, last_col).value):
        depth_col_idx = last_col
        next_status_col = last_col  # Insert before Depth
    else:
        next_status_col = last_col + 1
    
    # Add new STATUS and TIME headers
    from openpyxl.utils import get_column_letter
    
    # Count existing checks (pairs of STATUS/TIME columns)
    check_num = 1
    for col in range(3, last_col + 1, 2):
        if ws.cell(1, col).value and "STATUS" in str(ws.cell(1, col).value):
            check_num += 1
    
    status_col = get_column_letter(next_status_col)
    time_col = get_column_letter(next_status_col + 1)
    
    # Headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws[f'{status_col}1'] = f'STATUS (CHECK {check_num})'
    ws[f'{status_col}1'].fill = header_fill
    ws[f'{status_col}1'].font = header_font
    ws[f'{status_col}1'].alignment = Alignment(horizontal='center')
    
    ws[f'{time_col}1'] = f'TIME (CHECK {check_num})'
    ws[f'{time_col}1'].fill = header_fill
    ws[f'{time_col}1'].font = header_font
    ws[f'{time_col}1'].alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions[status_col].width = 12
    ws.column_dimensions[time_col].width = 12
    
    # Define colors for status
    checked_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    skipped_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
    
    # Update each market row
    for idx, (symbol, target) in enumerate(PAIRS, start=2):
        if symbol in results_dict:
            result = results_dict[symbol]
            status = result['status']
            
            # Status cell
            status_cell = ws[f'{status_col}{idx}']
            status_cell.value = status
            status_cell.alignment = Alignment(horizontal='center')
            
            # Color based on status
            if status == 'CHECKED':
                status_cell.fill = checked_fill
            elif status == 'WARNING':
                status_cell.fill = warning_fill
            else:  # SKIPPED
                status_cell.fill = skipped_fill
            
            # Time cell
            time_cell = ws[f'{time_col}{idx}']
            time_cell.value = check_time
            time_cell.alignment = Alignment(horizontal='center')
            
            # Update or create Depth column (always last)
            depth_display = f"{result['depth_1.25x']} / {result['depth_1.5x']}"
            
            if depth_col_idx:
                # Update existing Depth column
                depth_col = get_column_letter(depth_col_idx)
            else:
                # Create new Depth column
                depth_col_idx = next_status_col + 2
                depth_col = get_column_letter(depth_col_idx)
                ws[f'{depth_col}1'] = 'Depth (1.25x / 1.5x)'
                ws[f'{depth_col}1'].fill = header_fill
                ws[f'{depth_col}1'].font = header_font
                ws[f'{depth_col}1'].alignment = Alignment(horizontal='center')
                ws.column_dimensions[depth_col].width = 20
            
            depth_cell = ws[f'{depth_col}{idx}']
            depth_cell.value = depth_display
            depth_cell.alignment = Alignment(horizontal='center')
        else:
            # Mark as SKIPPED if not in results
            status_cell = ws[f'{status_col}{idx}']
            status_cell.value = 'SKIPPED'
            status_cell.fill = skipped_fill
            status_cell.alignment = Alignment(horizontal='center')
            
            time_cell = ws[f'{time_col}{idx}']
            time_cell.value = check_time
            time_cell.alignment = Alignment(horizontal='center')
    
    wb.save(log_path)
    print(f"✅ Daily log updated: {log_path}")

# --- Main Execution ---

def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    state = load_state()
    driver = init_driver()
    final_results = []
    excel_results = {}  # For Excel logging
    
    send_telegram(f"🔄 <b>Starting Hourly Check</b>\nPairs: {len(PAIRS)}")

    try:
        for symbol, target in PAIRS:
            attempt = 0
            success = False
            
            while attempt < MAX_ATTEMPTS_PER_PAIR and not success:
                attempt += 1
                try:
                    driver.get(f"{BASE_URL}{symbol}")
                    wait = WebDriverWait(driver, 15)
                    selector = ".newTrade-depth-block.depath-index-container"
                    element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    wait.until(lambda d: "Spread" in element.text and any(c.isdigit() for c in element.text))
                    
                    asks_df, bids_df, curr_spread = parse_orderbook(element.text)
                    if curr_spread is None: raise ValueError("No spread data")

                    # Calculations
                    diff = ((curr_spread - target) / target) * 100
                    is_poor = (diff > 100 or diff < -40)
                    dws = calculate_dws(asks_df, bids_df)
                    depth_25 = calculate_liquidity_depth(asks_df, bids_df, curr_spread * 1.25)
                    depth_50 = calculate_liquidity_depth(asks_df, bids_df, curr_spread * 1.5)

                    # State Logic
                    p_state = state.get(symbol, {"consecutive": 0, "last_alert": None, "start_time": None})
                    prev_status = "Warning" if p_state["consecutive"] > 0 else "Healthy"
                    
                    if is_poor:
                        p_state["consecutive"] += 1
                        if not p_state["start_time"]: 
                            p_state["start_time"] = get_nigerian_time().isoformat()
                    else:
                        p_state["consecutive"] = 0
                        p_state["start_time"] = None
                        p_state["last_alert"] = None

                    state[symbol] = p_state

                    # Alert Logic
                    if p_state["consecutive"] >= ALERT_THRESHOLD_CYCLES:
                        last_alert = p_state.get("last_alert")
                        cooldown_ok = True
                        if last_alert:
                            last_alert_time = datetime.fromisoformat(last_alert)
                            if last_alert_time.tzinfo is None:
                                last_alert_time = last_alert_time.replace(tzinfo=NIGERIAN_TZ)
                            if get_nigerian_time() - last_alert_time < timedelta(minutes=ALERT_COOLDOWN_MINUTES):
                                cooldown_ok = False
                        
                        if cooldown_ok:
                            alert_msg = f"⚠️ <b>ALERT: {symbol}</b>\nSpread: {curr_spread}%\nDiff: {diff:+.2f}%\nStrikes: {p_state['consecutive']}\nDepth 25%: {format_depth(depth_25)}"
                            send_telegram(alert_msg)
                            p_state["last_alert"] = get_nigerian_time().isoformat()

                    # Store results for CSV (backward compatibility)
                    final_results.append({
                        'timestamp': get_nigerian_time().strftime('%Y-%m-%d %H:%M:%S'),
                        'symbol': symbol, 'status': 'Warning' if is_poor else 'Healthy',
                        'strikes': p_state["consecutive"], 'current_spread': curr_spread,
                        'target_spread': target, 'percent_diff': round(diff, 2),
                        'dws': round(dws, 4), 'depth_1.25x': format_depth(depth_25), 'depth_1.5x': format_depth(depth_50)
                    })
                    
                    # Store results for Excel
                    excel_results[symbol] = {
                        'status': 'WARNING' if is_poor else 'CHECKED',
                        'depth_1.25x': format_depth(depth_25),
                        'depth_1.5x': format_depth(depth_50)
                    }
                    
                    success = True
                except Exception as e:
                    if attempt == MAX_ATTEMPTS_PER_PAIR:
                        # Mark as SKIPPED in Excel
                        excel_results[symbol] = {
                            'status': 'SKIPPED',
                            'depth_1.25x': 'N/A',
                            'depth_1.5x': 'N/A'
                        }
                        print(f"❌ Failed to scrape {symbol}: {str(e)}")
                    time.sleep(2)

        # Save latest.csv (backward compatibility)
        if final_results:
            new_df = pd.DataFrame(final_results)
            new_df.to_csv(os.path.join(DATA_DIR, "latest.csv"), index=False)
            print(f"✅ Data saved to latest.csv")
        
        # Update Excel daily log
        update_daily_log(excel_results)
        
        save_state(state)
        
        # Build completion message
        warnings = [r for r in final_results if r['status'] == 'Warning']
        total_pairs = len(final_results)
        skipped_count = len([r for r in excel_results.values() if r['status'] == 'SKIPPED'])
        
        completion_msg = f"✅ <b>Check Complete</b>\n"
        completion_msg += f"Total Pairs: {total_pairs}\n"
        completion_msg += f"Warnings: {len(warnings)}\n"
        completion_msg += f"Skipped: {skipped_count}\n"
        
        if warnings:
            completion_msg += f"\n<b>⚠️ Markets with Warnings:</b>\n"
            for w in warnings[:5]:  # Limit to first 5 to avoid long messages
                completion_msg += f"\n<b>{w['symbol']}</b>\n"
                completion_msg += f"  Spread: {w['current_spread']}% (Target: {w['target_spread']}%)\n"
                completion_msg += f"  Diff: {w['percent_diff']:+.2f}%\n"
                completion_msg += f"  Strikes: {w['strikes']}\n"
            
            if len(warnings) > 5:
                completion_msg += f"\n... and {len(warnings) - 5} more warnings"
        
        send_telegram(completion_msg)

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
